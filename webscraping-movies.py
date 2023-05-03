
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2023/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

wb= xl.Workbook()

ws = wb.active

ws.title = 'Box Office Report'

ws['A1'] = 'No.'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Release Date'
ws['D1'] = 'Gross'
ws['E1'] = 'Total Gross'
ws['F1'] = '% of Total Gross'
table_rows = soup.findAll('tr')

for x in range(1,6):
    td = table_rows[x].findAll('td')
    rank = td[0].text
    name = td[1].text
    gross = int(td[5].text.replace(',','').replace('$',''))
    total_gross = int(td[7].text.replace(',','').replace('$',''))
    date = td[8].text

    percent_gross = round((gross/total_gross) * 100,2)

    ws['A' + str(x+1)] = rank
    ws['B' + str(x+1)] = name
    ws['C' + str(x+1)] = date
    ws['D' + str(x+1)] = gross
    ws['E' + str(x+1)] = total_gross
    ws['F' + str(x+1)] = str(percent_gross)+'%'

header_font = Font(size = 16, bold = True)

for cell in ws[1:1]:
    cell.font = header_font


wb.save('Box Office Report.xlsx')