import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1, title = 'Second Sheet')

ws['A1'] = 'Invoice'
# ws['A1'].font = Font(name= 'Times New Roman', size =24, bold=True, italic=False)
myfont = Font(name= 'Times New Roman', size =24, bold=True, italic=False)
ws['A1'].font = myfont

ws['A2'] = 'Tires'
ws['A3']= 'Brakes'
ws['A4']= 'Alighment'

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = myfont
ws['B8'] = '=SUM(B2:B4)'

# ws.column_dimensions['A'] = 25

ws.merge_cells('A1:B1')

#WORK IN CLASS
produce = xl.load_workbook('ProduceReport.xlsx')
produce_ws = produce['ProduceReport']
ws2 = wb['Second Sheet']


ws2['A1']= 'Produce'
ws2['B1']= 'Cost per Pound'
ws2['C1']= 'Amount Sold'
ws2['D1']= 'Total'

write_row = 2
write_colA = 1
write_colB = 2
write_colC = 3
write_colD = 4

for currentrow in produce_ws.iter_rows(min_row=2, max_row=produce_ws.max_row, max_col= produce_ws.max_column):
    produces = currentrow[0].value
    cost = float(currentrow[1].value)
    amount_sold = float(currentrow[2].value)
    total = float(currentrow[3].value)

    ws2.cell(write_row,write_colA).value = produces
    ws2.cell(write_row,write_colB).value = cost
    ws2.cell(write_row,write_colC).value = amount_sold
    ws2.cell(write_row, write_colD).value = total
    
    write_row +=1

summary_row = write_row+1
ws2['B'+str(summary_row)] = 'Total' 
ws2['B'+str(summary_row)].font = Font(size=16, bold = True)

ws2['C'+str(summary_row)] = '=SUM(C2:C'+str(write_row)+')'
                            # same as saying =SUM(C2:C42)
ws2['D'+str(summary_row)] = '=SUM(D2:D'+str(write_row)+')'

summary_row +=1
ws2['B'+str(summary_row)] = 'Average' 
ws2['B'+str(summary_row)].font = Font(size=16, bold = True)

ws2['C'+str(summary_row)] = '=AVERAGE(C2:C'+str(write_row)+')'
                            
ws2['D'+str(summary_row)] = '=AVERAGE(D2:D'+str(write_row)+')'

ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 15

for cell in ws2['C:C']:
    cell.number_format = '#,##0'

for cell in ws2['D:D']:
    cell.number_format = u'"$"#,##0.00'




wb.save('PythontoExcel.xlsx')