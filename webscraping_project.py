from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font


url = 'https://www.coingecko.com/'


		
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

req = Request(url, headers= headers)

webpage = urlopen(req).read()

soup = BeautifulSoup(webpage, 'html.parser')

wb= xl.Workbook()

ws = wb.active

ws.title = 'Cryptocurrencies Report'

ws['A1'] = 'Name'
ws['B1'] = 'Symbol'
ws['C1'] = 'Current price'
ws['D1'] = '% change in 24h'
table_rows = soup.findAll('tr')

for x in range(1,6):
    td = table_rows[x].findAll('td')
    name = td[2]
    find_name = name.findAll('span') 
    actual_name = find_name[0].text.strip('\n')
    symbol = find_name[1].text.strip('\n')
    price = float(td[3].text.replace(',','').replace('$','').strip('\n'))
    percent_change = float(td[5].text.replace(',','').replace('%','').strip('\n'))
    
    ws['A' + str(x+1)] = actual_name
    ws['B' + str(x+1)] = symbol
    ws['C' + str(x+1)] = price
    ws['D' + str(x+1)] = str(percent_change)+'%'

    
for cell in ws['C:C']:
    cell.number_format = u'"$"#,##0.00'

ws.column_dimensions['C'].width = 15

header_font = Font(size = 16, bold = True)

for cell in ws[1:1]:
    cell.font = header_font


import keys
from twilio.rest import Client

Twilionumber = '+12148335896'

mycellphone = '+12542142839'

if name == 'Bitcoin' or 'Ethereum':
    if percent_change >= 0.02:
        text = 'Bitcoin or Ethereum have increased its price by $5 or more'
    
    textmessage = Client.messages.create(to=mycellphone, from_=Twilionumber, 
                                     body= text)     


wb.save('Cryptocurrencies Report.xlsx')








